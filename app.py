import streamlit as st
from google import genai
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import PIL.Image
import json
import time
import io
import zipfile
from datetime import datetime, timezone, timedelta

st.set_page_config(page_title="Registro de Medidores", page_icon="📊", layout="centered")
st.title("📊 Registro de Medidores de Gas")
st.markdown("Sube hasta 12 fotos de medidores y genera el reporte automáticamente.")

API_KEYS = [
    st.secrets["GEMINI_API_KEY_1"],
    st.secrets["GEMINI_API_KEY_2"],
    st.secrets["GEMINI_API_KEY_3"],
]

OPERARIOS = ["Joseph Erik Abanto Guerra", "Marco David Rodríguez Valencia"]

OBSERVACIONES = [
    "NINGUNA",
    "PRESENCIA DE CORROSIÓN SOBRE MEDIDOR",
    "PRESENCIA DE GOLPE SOBRE MEDIDOR",
    "PRESENCIA DE CONTAMINANTE SOBRE MEDIDOR",
    "PRESENCIA DE RAYADURA SOBRE MEDIDOR",
    "ODÓMETRO SE DETIENE",
    "ODÓMETRO NO GIRA",
    "AUSENCIA DE PRECINTO DE SEGURIDAD",
    "PRECINTO DE SEGURIDAD ROTO",
    "PRESENCIA DE CONTAMINANTES SOBRE LUNETA",
    "PRESENCIA DE FISURA SOBRE LUNETA",
    "PRESENCIA DE RAYADURA SOBRE LUNETA",
    "PRESENCIA DE CONTAMINANTE SOBRE ROSCA",
    "PRESENCIA DE CORROSIÓN SOBRE ROSCA",
    "PRESENCIA DE GOLPE SOBRE ROSCA",
    "SELLOS DE SEGURIDAD DE ODÓMETRO ROTOS",
    "AUSENCIA DE TORNILLOS DE SEGURIDAD",
]

prompt = """Analiza esta imagen de un medidor de gas con extremo detalle. Tu misión es extraer TODOS los datos visibles, especialmente el número del precinto de seguridad.
Extrae exactamente estos datos y devuélvelos SOLO en formato JSON sin texto adicional:
{
  "marca": "",
  "modelo": "",
  "serie_medidor": "",
  "registro": "",
  "unidad": "",
  "volumen_ciclico": "",
  "serie_precinto": ""
}

Instrucciones detalladas:
- marca: nombre del fabricante (ej: METREX)
- modelo: capacidad del medidor, siempre con punto (ej: G1.6 no G1,6 ni G16)
- serie_medidor: numero de serie del medidor (ej: 3809043)
- registro: SOLAMENTE los 5 primeros dígitos del contador (los enteros que están ANTES de la coma). NO incluyas la coma ni los dígitos rojos. Ejemplos: si ves "00123,422" devuelve "00123". Si ves "00954,095" devuelve "00954".
- unidad: unidad de medida del registro (ej: m3)
- volumen_ciclico: valor V indicado en la placa (ej: 0.7 dm3)

INSTRUCCIONES CRÍTICAS PARA EL PRECINTO:
El precinto de seguridad es una etiqueta pequeña colgante que cuelga o está pegada al medidor. Es un dato OBLIGATORIO que SIEMPRE debes encontrar.

REGLAS PARA LEER EL PRECINTO:
1. Busca en TODA la imagen — no solo en una zona.
2. La etiqueta puede estar en CUALQUIER ÁNGULO: 0°, 45°, 90°, 135°, 180°, 225°, 270°, 315°.
3. Si ves caracteres como "GUM", "8YL", letras invertidas, intenta leer mentalmente rotando la imagen.
4. El formato OBLIGATORIO es: letra G mayúscula + 7 dígitos numéricos (ej: G0454825, G0457243).
5. Si después de TODAS las rotaciones no es claramente visible, solo entonces pon null.

Si no puedes leer algun otro dato, pon null."""

if "tabla" not in st.session_state:
    st.session_state.tabla = []
if "fotos_bytes" not in st.session_state:
    st.session_state.fotos_bytes = []
if "procesado" not in st.session_state:
    st.session_state.procesado = False
if "operario_guardado" not in st.session_state:
    st.session_state.operario_guardado = ""
if "fecha_guardada" not in st.session_state:
    st.session_state.fecha_guardada = ""
if "lote_guardado" not in st.session_state:
    st.session_state.lote_guardado = ""
if "observaciones_lote" not in st.session_state:
    st.session_state.observaciones_lote = []
if "registros_finales" not in st.session_state:
    st.session_state.registros_finales = []
if "lotes_eliminados" not in st.session_state:
    st.session_state.lotes_eliminados = []
if "contador_lote_dia" not in st.session_state:
    st.session_state.contador_lote_dia = {}

def generar_lote(fecha_dt):
    fecha_clave = fecha_dt.strftime("%Y-%m-%d")
    if fecha_clave not in st.session_state.contador_lote_dia:
        st.session_state.contador_lote_dia[fecha_clave] = 0
    st.session_state.contador_lote_dia[fecha_clave] += 1
    qv = st.session_state.contador_lote_dia[fecha_clave]
    return f"QV{qv}-VP-{fecha_clave}"

def registro_inicial_de(datos):
    """Devuelve el registro inicial de un medidor como número entero, sin ceros a la izquierda."""
    return int(str(datos["registro"]).lstrip("0") or "0")

def procesar_imagen(imagen):
    ultimo_error = None
    for key_idx, api_key in enumerate(API_KEYS):
        try:
            client = genai.Client(api_key=api_key)
            respuesta = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=[prompt, imagen]
            )
            texto = respuesta.text.strip().replace("```json","").replace("```","")
            return json.loads(texto), None
        except Exception as e:
            ultimo_error = str(e)
            if key_idx < len(API_KEYS) - 1:
                time.sleep(3)
                continue
            else:
                return None, ultimo_error

def estilo(ws, celda, valor, negrita=False, tam=10, color_texto="000000",
           color_fondo=None, alineacion="center", borde=None, alto=None, italica=False):
    if isinstance(celda, str):
        c = ws[celda]
    else:
        c = celda
    try:
        c.value = valor
    except AttributeError:
        return None
    c.font = Font(bold=negrita, size=tam, color=color_texto, italic=italica)
    c.alignment = Alignment(horizontal=alineacion, vertical="center", wrap_text=True)
    if color_fondo:
        c.fill = PatternFill("solid", fgColor=color_fondo)
    if borde:
        c.border = borde
    return c

def generar_excel(tabla, fotos_bytes, operario, fecha, lote, nombres_imagenes, observaciones, registros_finales):
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro"

    AZUL = "17375E"
    AZUL_CLARO = "D9EAF7"
    VERDE = "2E8B57"
    VERDE_CLARO = "EAF7EE"
    ROJO = "C0392B"
    GRIS = "F4F6F7"
    BLANCO = "FFFFFF"
    AMARILLO = "FCF3CF"
    NARANJA = "E67E22"
    AZUL_LINK = "1E40AF"

    borde_fino = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    borde_medio = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )

    total = len(tabla)
    correctos = total

    anchos = [5, 10, 10, 16, 11, 10, 10, 16, 16, 16, 12, 10, 35, 5, 14, 5, 12, 8, 2.11]
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a

    # FILA 1: TÍTULO
    ws.merge_cells("A1:S1")
    estilo(ws, "A1", "LABORATORIO DE MEDIDORES DE GAS",
           negrita=True, tam=18, color_texto=BLANCO, color_fondo=AZUL, borde=borde_medio)
    ws.row_dimensions[1].height = 35

    # FILA 2: SUBTÍTULO
    ws.merge_cells("A2:S2")
    estilo(ws, "A2", "Reporte automático de inspección y registro de medidores",
           tam=10, italica=True, color_texto="404040", color_fondo=GRIS, borde=borde_fino)
    ws.row_dimensions[2].height = 18

    # FILA 3: DATOS GENERALES
    ws.merge_cells("A3:D3")
    estilo(ws, "A3", f"👤 Operario: {operario}", negrita=True,
           color_fondo=AZUL_CLARO, borde=borde_fino, alineacion="left")
    ws.merge_cells("E3:L3")
    estilo(ws, "E3", f"📅 Fecha: {fecha}", negrita=True,
           color_fondo=AZUL_CLARO, borde=borde_fino)
    ws.merge_cells("M3:S3")
    estilo(ws, "M3", f"🏷 Lote: {lote}", negrita=True,
           color_texto=BLANCO, color_fondo=AZUL, borde=borde_fino)
    ws.row_dimensions[3].height = 22

    ws.row_dimensions[4].height = 6

    # FILAS 5-6: TARJETAS
    marca_tarjeta = (tabla[0].get("marca") or "METREX") if tabla else "-"
    modelo_tarjeta = (tabla[0].get("modelo") or "G1.6") if tabla else "-"
    estado = "APROBADO"

    tarjetas = [
        ("A5", "D6", "TOTAL DE MEDIDORES", str(total)),
        ("E5", "H6", "MARCA DEL LOTE", marca_tarjeta),
        ("I5", "N6", "MODELO", modelo_tarjeta),
        ("O5", "S6", "ESTADO DEL LOTE", estado),
    ]

    for inicio, fin, titulo, valor in tarjetas:
        ws.merge_cells(f"{inicio}:{fin}")
        estilo(ws, inicio, f"{titulo}\n{valor}",
               negrita=True, tam=11, color_texto=BLANCO, color_fondo=AZUL, borde=borde_medio)
        ws[inicio].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    for r in [5, 6]:
        ws.row_dimensions[r].height = 22

    ws.row_dimensions[7].height = 6

    # FILA 8: ENCABEZADO TABLA
    ws.merge_cells("A8:M8")
    estilo(ws, "A8", "RESUMEN DE MEDIDORES DEL LOTE", negrita=True, tam=11,
           color_texto=BLANCO, color_fondo=AZUL, borde=borde_fino)
    ws.row_dimensions[8].height = 22

    # FILA 9: SUBENCABEZADOS
    cols_tabla = ["#", "Marca", "Modelo", "Nro. Serie Medidor", "Vol. Cíclico",
                  "Tipo", "Color", "Nro. Serie Precinto", "Registro Inicial (m³)",
                  "Registro Final (m³)", "Diferencia (m³)", "Foto", "Observación"]
    for i, h in enumerate(cols_tabla):
        c = ws.cell(row=9, column=i+1)
        estilo(ws, c, h, negrita=True, tam=9, color_texto=BLANCO,
               color_fondo=ROJO, borde=borde_fino)
    ws.row_dimensions[9].height = 26.4

    # FILAS DE DATOS
    MODELO_DEFAULT = "G1.6"
    MARCA_DEFAULT = "METREX"

    for idx, datos in enumerate(tabla):
        fila = 10 + idx
        color_fila = GRIS if idx % 2 == 0 else BLANCO
        vol_normalizado = "0.7 dm³"
        unidad_normalizada = "m³"
        marca = datos.get("marca") or MARCA_DEFAULT
        modelo = datos.get("modelo") or MODELO_DEFAULT
        obs = observaciones[idx] if idx < len(observaciones) else "NINGUNA"
        # Limpiar ceros a la izquierda del registro
        registro_limpio = str(datos["registro"]).lstrip("0") or "0"
        registro_inicial_num = registro_inicial_de(datos)
        registro_final_num = registros_finales[idx] if idx < len(registros_finales) else registro_inicial_num
        diferencia_num = registro_final_num - registro_inicial_num
        valores = [
            idx+1, marca, modelo, datos["serie_medidor"],
            vol_normalizado, "Circular", "Verde",
            datos["serie_precinto"] or "N/D",
            f"{registro_limpio} {unidad_normalizada}",
            f"{registro_final_num} {unidad_normalizada}",
            f"{diferencia_num} {unidad_normalizada}",
            "", obs
        ]
        for col, val in enumerate(valores, 1):
            c = ws.cell(row=fila, column=col)
            color = color_fila
            if col == 11 and diferencia_num < 0:
                color = "FDEDEC"
            if col == 13 and obs != "NINGUNA":
                color = "FDEDEC"
            estilo(ws, c, val, tam=9, color_fondo=color, borde=borde_fino)
            if col == 13:
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        # Hipervínculo
        nombre_img = nombres_imagenes[idx]
        c_link = ws.cell(row=fila, column=12)
        c_link.value = f"📷 Ver foto {idx+1}"
        c_link.hyperlink = f"imagenes/{nombre_img}"
        c_link.font = Font(bold=True, size=9, color=AZUL_LINK, underline="single")
        c_link.alignment = Alignment(horizontal="center", vertical="center")
        c_link.fill = PatternFill("solid", fgColor=VERDE_CLARO)
        c_link.border = borde_fino
        ws.row_dimensions[fila].height = 16

    # PANEL DETALLES
    ws.merge_cells("N8:S8")
    estilo(ws, "N8", "DETALLES DEL LOTE", negrita=True, tam=11,
           color_texto=BLANCO, color_fondo=AZUL, borde=borde_fino)

    detalles = [
        ("Lote:", lote),
        ("Fecha y hora:", fecha),
        ("Operario:", operario),
        ("Marca:", marca_tarjeta),
        ("Modelo:", modelo_tarjeta),
        ("Tipo:", "Circular"),
        ("Vol. Cíclico:", "0.7 dm³ (fijo)"),
        ("Color:", "Verde"),
        ("Total medidores:", str(total)),
        ("Supervisados:", f"{correctos} (100%)"),
    ]

    for i, (clave, valor) in enumerate(detalles):
        fila_d = 9 + i
        ws.merge_cells(f"N{fila_d}:O{fila_d}")
        estilo(ws, f"N{fila_d}", clave, negrita=True, tam=8,
               color_fondo=AZUL_CLARO, borde=borde_fino, alineacion="left")
        ws.merge_cells(f"P{fila_d}:S{fila_d}")
        estilo(ws, f"P{fila_d}", valor, tam=8,
               color_fondo=BLANCO, borde=borde_fino, alineacion="left")
        ws.row_dimensions[fila_d].height = 16

    cert_fila = 9 + len(detalles) + 1
    ws.merge_cells(f"N{cert_fila}:S{cert_fila+2}")
    estilo(ws, f"N{cert_fila}",
           f"Certifico que el proceso fue supervisado correctamente\n— {operario} —\n{fecha}",
           negrita=True, italica=True, tam=8,
           color_fondo=VERDE_CLARO, borde=borde_medio, alineacion="center")
    ws.row_dimensions[cert_fila].height = 18
    ws.row_dimensions[cert_fila+1].height = 18
    ws.row_dimensions[cert_fila+2].height = 18

    # ESTADÍSTICAS
    est_fila = cert_fila + 4
    ws.merge_cells(f"N{est_fila}:S{est_fila}")
    estilo(ws, f"N{est_fila}", "RESUMEN DE OBSERVACIONES", negrita=True, tam=10,
           color_texto=BLANCO, color_fondo=VERDE, borde=borde_fino)
    ws.row_dimensions[est_fila].height = 20

    sin_obs = observaciones.count("NINGUNA")
    con_obs = total - sin_obs

    stats = [
        ("Sin observación", sin_obs, VERDE_CLARO),
        ("Con observación", con_obs, AMARILLO),
    ]
    for i, (nombre, valor, color) in enumerate(stats):
        fr = est_fila + 1 + i
        ws.merge_cells(f"N{fr}:P{fr}")
        estilo(ws, f"N{fr}", nombre, tam=9, color_fondo=color, borde=borde_fino)
        ws.merge_cells(f"Q{fr}:S{fr}")
        pct = f"{valor} ({int(valor/total*100) if total else 0}%)"
        estilo(ws, f"Q{fr}", pct, negrita=True, tam=9, color_fondo=color, borde=borde_fino)
        ws.row_dimensions[fr].height = 16

    # INFO FOTOS
    info_fila = 10 + total + 2
    ws.merge_cells(f"A{info_fila}:M{info_fila}")
    estilo(ws, f"A{info_fila}",
           "💡 Las fotos están en la carpeta 'imagenes/'. Haz clic en cada enlace de la columna Foto para abrirlas.",
           tam=10, italica=True, color_texto="404040", color_fondo=AMARILLO, borde=borde_fino)
    ws.row_dimensions[info_fila].height = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_zip(tabla, fotos_bytes, operario, fecha, lote, observaciones, registros_finales):
    nombres_imagenes = []
    for idx, datos in enumerate(tabla):
        serie = datos.get("serie_medidor", f"med{idx+1}")
        nombre = f"{str(idx+1).zfill(2)}_{serie}.png"
        nombres_imagenes.append(nombre)

    excel_buffer = generar_excel(tabla, fotos_bytes, operario, fecha, lote, nombres_imagenes, observaciones, registros_finales)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        nombre_excel = f"Reporte_{lote}.xlsx"
        zf.writestr(nombre_excel, excel_buffer.read())
        for idx, (fb, nombre_img) in enumerate(zip(fotos_bytes, nombres_imagenes)):
            try:
                img = PIL.Image.open(io.BytesIO(fb))
                img_buffer = io.BytesIO()
                img.save(img_buffer, format="PNG")
                img_buffer.seek(0)
                zf.writestr(f"imagenes/{nombre_img}", img_buffer.read())
            except Exception:
                pass

    zip_buffer.seek(0)
    return zip_buffer

# INTERFAZ
st.markdown("---")
operario = st.selectbox("Selecciona el operario:", OPERARIOS)
peru = timezone(timedelta(hours=-5))
ahora = datetime.now(peru)
fecha = ahora.strftime("%d/%m/%Y %H:%M")
st.info(f"Fecha y hora: {fecha}")

st.markdown("---")
fotos = st.file_uploader(
    "Sube las fotos de los medidores (hasta 12)",
    type=["jpg", "jpeg", "png"],
    accept_multiple_files=True
)

if fotos:
    st.info(f"{len(fotos)} foto(s) seleccionada(s)")
    if len(fotos) > 12:
        st.warning(f"⚠️ Subiste {len(fotos)} fotos, pero solo se procesarán las primeras 12. Las {len(fotos) - 12} restantes serán ignoradas.")

    if st.button("Procesar medidores", type="primary"):
        st.session_state.tabla = []
        st.session_state.fotos_bytes = []
        st.session_state.procesado = False
        st.session_state.operario_guardado = operario
        st.session_state.fecha_guardada = fecha
        st.session_state.lote_guardado = generar_lote(ahora)
        st.session_state.observaciones_lote = []
        st.session_state.registros_finales = []

        progress = st.progress(0)
        status = st.empty()

        for i, foto in enumerate(fotos[:12]):
            status.text(f"Procesando medidor {i+1} de {len(fotos)}...")
            foto.seek(0)
            foto_bytes = foto.read()
            st.session_state.fotos_bytes.append(foto_bytes)
            imagen = PIL.Image.open(io.BytesIO(foto_bytes))
            datos, error = procesar_imagen(imagen)

            if datos:
                # Limpiar ceros a la izquierda del registro
                if datos.get("registro"):
                    datos["registro"] = str(datos["registro"]).lstrip("0") or "0"
                st.session_state.tabla.append(datos)
                st.session_state.observaciones_lote.append("")
                st.session_state.registros_finales.append(registro_inicial_de(datos))
                st.success(f"Medidor {i+1}: Serie {datos['serie_medidor']} | Registro {datos['registro']} {datos['unidad']} | Precinto {datos['serie_precinto']}")
            else:
                st.error(f"No se pudo procesar el medidor {i+1}. Motivo: {error or 'desconocido'}")

            progress.progress((i+1) / len(fotos))
            time.sleep(5)

        st.session_state.procesado = True
        status.text("Procesamiento completado.")

if st.session_state.procesado and st.session_state.tabla:
    tabla = st.session_state.tabla
    fotos_bytes = st.session_state.fotos_bytes
    operario = st.session_state.operario_guardado
    fecha = st.session_state.fecha_guardada
    lote = st.session_state.lote_guardado

    st.markdown("---")
    st.subheader(f"Lote: {lote}")
    st.subheader("Tabla completa del lote")

    st.markdown("**Ingresa el registro final y selecciona una observación para cada medidor:**")

    while len(st.session_state.registros_finales) < len(tabla):
        st.session_state.registros_finales.append(0)

    for i, d in enumerate(tabla):
        registro_inicial_num = registro_inicial_de(d)

        col1, col2, col3 = st.columns([2, 2, 2])
        with col1:
            st.markdown(f"**#{i+1}** — Serie: `{d['serie_medidor']}`")
            st.caption(f"Registro inicial: {registro_inicial_num} m³")
        with col2:
            seleccion = st.selectbox(
                f"Observación medidor {i+1}",
                options=[""] + OBSERVACIONES,
                key=f"obs_{i}",
                label_visibility="collapsed"
            )
            if i < len(st.session_state.observaciones_lote):
                st.session_state.observaciones_lote[i] = seleccion
        with col3:
            reg_final = st.number_input(
                f"Registro final medidor {i+1}",
                min_value=0,
                value=int(st.session_state.registros_finales[i]),
                step=1,
                key=f"regfinal_{i}",
                label_visibility="collapsed"
            )
            st.session_state.registros_finales[i] = reg_final
            diferencia = reg_final - registro_inicial_num
            if diferencia < 0:
                st.caption(f"⚠️ Diferencia: {diferencia} m³ (revisar lectura)")
            else:
                st.caption(f"Diferencia: {diferencia} m³")

    st.markdown("---")

    with st.expander("🗑️ Eliminar este lote"):
        st.caption("Si este lote no debe continuar (error de captura, prueba anulada, etc.), puedes eliminarlo. Se descartarán los datos y solo quedará registrado el número de lote y el motivo.")
        motivo_eliminacion = st.text_area(
            "Motivo de la eliminación",
            key="motivo_elim",
            placeholder="Ej: Se repitió el lote por error en la carga de fotos"
        )
        confirmar_eliminacion = st.checkbox(f"Confirmo que quiero eliminar el lote {lote}. Esta acción no se puede deshacer.")
        if st.button("🗑️ Eliminar lote", type="secondary", disabled=not confirmar_eliminacion):
            if motivo_eliminacion.strip():
                st.session_state.lotes_eliminados.append({
                    "lote": lote,
                    "motivo": motivo_eliminacion.strip()
                })
                st.session_state.tabla = []
                st.session_state.fotos_bytes = []
                st.session_state.procesado = False
                st.session_state.observaciones_lote = []
                st.session_state.registros_finales = []
                st.success(f"Lote {lote} eliminado. Motivo registrado.")
                st.rerun()
            else:
                st.warning("Debes escribir un motivo antes de eliminar el lote.")

    st.markdown("---")

    todas_llenas = all(obs != "" for obs in st.session_state.observaciones_lote)
    if not todas_llenas:
        st.warning("⚠️ Debes seleccionar una observación para cada medidor antes de continuar.")

    confirmado = st.checkbox(f"Confirmo que el proceso fue supervisado correctamente por {operario}")

    if confirmado and todas_llenas:
        nombre_archivo = f"MetriLab_{lote}.zip"
        zip_file = generar_zip(tabla, fotos_bytes, operario, fecha, lote, st.session_state.observaciones_lote, st.session_state.registros_finales)

        st.download_button(
            label="📥 Descargar reporte completo (ZIP)",
            data=zip_file,
            file_name=nombre_archivo,
            mime="application/zip"
        )
        st.info("El ZIP contiene el Excel y la carpeta de imágenes. Descomprime para usar los enlaces.")
    elif not confirmado:
        st.warning("Debes confirmar la supervisión antes de descargar el reporte.")

if st.session_state.lotes_eliminados:
    st.markdown("---")
    st.subheader("🗑️ Lotes eliminados en esta sesión")
    for item in st.session_state.lotes_eliminados:
        st.markdown(f"- **{item['lote']}** — Motivo: {item['motivo']}")
